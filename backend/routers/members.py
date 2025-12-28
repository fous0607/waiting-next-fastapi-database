from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Response
from sqlalchemy.orm import Session
from sqlalchemy import or_, func, desc, and_
from typing import List, Optional
import openpyxl
from io import BytesIO
from datetime import datetime, date, timedelta

from database import get_db
from models import Member, Store, WaitingList, StoreSettings
from schemas import (
    Member as MemberSchema,
    MemberCreate,
    MemberUpdate,
    MemberBulkCreate
)
from auth import get_current_store
from sse_manager import sse_manager
from core.logger import logger

router = APIRouter()

def check_member_uniqueness(db: Session, store: Store, phone: str = None, barcode: str = None, exclude_member_id: int = None):
    """
    회원 중복 체크 로직
    - Barcode: 전역적 유일성 체크
    - Phone: 매장/프랜차이즈 범위 내 중복 체크
    
    Returns: (conflict_type: str|None, existing_member: Member|None)
    conflict_type: 'barcode' or 'phone'
    """
    
    # 1. Barcode Check (Global Uniqueness)
    if barcode:
        query = db.query(Member).filter(Member.barcode == barcode)
        if exclude_member_id:
            query = query.filter(Member.id != exclude_member_id)
        existing = query.first()
        if existing:
            return "barcode", existing

    # 2. Phone Check (Scoped Uniqueness)
    if phone:
        # 프랜차이즈 설정 확인 (기본값 store)
        member_type = "store"
        if store.franchise:
            member_type = store.franchise.member_type
        
        query = db.query(Member)
        
        if member_type == "franchise":
            # 프랜차이즈 내 모든 매장 검색
            store_ids = db.query(Store.id).filter(Store.franchise_id == store.franchise_id).all()
            store_ids = [s[0] for s in store_ids]
            query = query.filter(Member.store_id.in_(store_ids))
        else:
            # 매장 내 검색
            query = query.filter(Member.store_id == store.id)
            
        # 핸드폰 번호 체크
        query = query.filter(Member.phone == phone)
        
        # 수정 시 자기 자신 제외
        if exclude_member_id:
            query = query.filter(Member.id != exclude_member_id)
            
        existing = query.first()
        if existing:
            return "phone", existing

    return None, None

@router.post("", response_model=MemberSchema)
async def create_member(
    member: MemberCreate,
    current_store: Store = Depends(get_current_store),
    db: Session = Depends(get_db)
):
    """회원 등록"""
    # 중복 확인
    conflict_type, existing = check_member_uniqueness(db, current_store, phone=member.phone, barcode=member.barcode)
    if existing:
        if conflict_type == "barcode":
            raise HTTPException(status_code=400, detail="이미 등록된 바코드입니다.")
        else:
            msg = "이미 등록된 핸드폰번호입니다."
            if current_store.franchise and current_store.franchise.member_type == "franchise":
                msg += " (프랜차이즈 통합 관리)"
            raise HTTPException(status_code=400, detail=msg)

    db_member = Member(**member.dict(), store_id=current_store.id)
    db.add(db_member)
    db.commit()
    db.refresh(db_member)

    # 대기 목록 동기화: 핸드폰 번호로 대기 중인 항목 찾아 member_id 연결
    active_waitings = db.query(WaitingList).filter(
        WaitingList.phone == db_member.phone,
        WaitingList.status == "waiting",
        WaitingList.store_id == current_store.id
    ).all()

    for w in active_waitings:
        w.member_id = db_member.id
        # w.name = db_member.name # 이름도 동기화 (선택적)
    
    if active_waitings:
        db.commit()
        
        # 1. 관리자(Admin)에게는 무조건 전송
        await sse_manager.broadcast(
            store_id=str(current_store.id),
            event_type="member_updated",
            data={
                "member_id": db_member.id,
                "name": db_member.name,
                "phone": db_member.phone
            },
            franchise_id=str(current_store.franchise_id) if current_store.franchise_id else None,
            target_role='admin'
        )

        # 2. 대기현황판(Board)에게는 설정이 켜져 있을 때만 전송
        settings = db.query(StoreSettings).filter(StoreSettings.store_id == current_store.id).first()
        should_broadcast_board = True
        if settings:
            use_board = getattr(settings, 'enable_waiting_board', True) 
            use_desk = getattr(settings, 'enable_reception_desk', True)
            if not use_board and not use_desk:
                should_broadcast_board = False
                logger.debug(f"Skipping member_updated broadcast to BOARD/RECEPTION (Board: disabled, Reception: disabled): store_id={current_store.id}")

        if should_broadcast_board:
            await sse_manager.broadcast(
                store_id=str(current_store.id),
                event_type="member_updated",
                data={
                    "member_id": db_member.id,
                    "name": db_member.name,
                    "phone": db_member.phone
                },
                franchise_id=None,
                target_role='board'
            )

    return db_member

@router.get("", response_model=List[MemberSchema])
async def get_members(
    skip: int = 0,
    limit: int = 100,
    search: Optional[str] = None,
    current_store: Store = Depends(get_current_store),
    db: Session = Depends(get_db)
):
    """회원 목록 조회"""
    query = db.query(Member).filter(Member.store_id == current_store.id)

    # 검색 조건 (이름 또는 핸드폰번호)
    # 검색 조건 (이름 또는 핸드폰번호)
    if search:
        # 검색어가 4자리 숫자인 경우: 핸드폰 뒷자리 검색으로 간주하여 endswith 사용
        if search.isdigit() and len(search) == 4:
            query = query.filter(
                or_(
                    Member.name.contains(search),
                    Member.phone.endswith(search),
                    Member.barcode == search
                )
            )
        # 그 외의 경우: 포함 여부로 검색 (+바코드 정확 일치)
        else:
            query = query.filter(
                or_(
                    Member.name.contains(search),
                    Member.phone.contains(search),
                    Member.barcode == search
                )
            )

    members = query.offset(skip).limit(limit).all()
    return members

@router.get("/{member_id}", response_model=MemberSchema)
async def get_member(
    member_id: int,
    current_store: Store = Depends(get_current_store),
    db: Session = Depends(get_db)
):
    """회원 상세 조회"""
    member = db.query(Member).filter(
        Member.id == member_id,
        Member.store_id == current_store.id
    ).first()

    if not member:
        raise HTTPException(status_code=404, detail="회원을 찾을 수 없습니다.")

    return member

@router.get("/phone/{phone}", response_model=MemberSchema)
async def get_member_by_phone(
    phone: str,
    current_store: Store = Depends(get_current_store),
    db: Session = Depends(get_db)
):
    """핸드폰번호로 회원 조회"""
    member = db.query(Member).filter(
        Member.phone == phone,
        Member.store_id == current_store.id
    ).first()

    if not member:
        raise HTTPException(status_code=404, detail="회원을 찾을 수 없습니다.")

    return member

@router.put("/{member_id}", response_model=MemberSchema)
async def update_member(
    member_id: int,
    member: MemberUpdate,
    current_store: Store = Depends(get_current_store),
    db: Session = Depends(get_db)
):
    """회원 정보 수정"""
    db_member = db.query(Member).filter(
        Member.id == member_id,
        Member.store_id == current_store.id
    ).first()

    if not db_member:
        raise HTTPException(status_code=404, detail="회원을 찾을 수 없습니다.")

    # 핸드폰번호/바코드 변경 시 중복 체크
    check_phone = member.phone if (member.phone and member.phone != db_member.phone) else None
    check_barcode = member.barcode if (member.barcode and member.barcode != db_member.barcode) else None

    if check_phone or check_barcode:
        conflict_type, existing = check_member_uniqueness(db, current_store, phone=check_phone, barcode=check_barcode, exclude_member_id=member_id)
        if existing:
            if conflict_type == "barcode":
                 raise HTTPException(status_code=400, detail="이미 등록된 바코드입니다.")
            else:
                msg = "이미 등록된 핸드폰번호입니다."
                if current_store.franchise and current_store.franchise.member_type == "franchise":
                    msg += " (프랜차이즈 통합 관리)"
                raise HTTPException(status_code=400, detail=msg)

    # 업데이트할 필드만 수정
    update_data = member.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_member, field, value)

    db.commit()
    db.refresh(db_member)

    # 1. 관리자(Admin)에게는 무조건 전송
    await sse_manager.broadcast(
        store_id=str(current_store.id),
        event_type="member_updated",
        data={
            "member_id": db_member.id,
            "name": db_member.name,
            "phone": db_member.phone
        },
        franchise_id=str(current_store.franchise_id) if current_store.franchise_id else None,
        target_role='admin'
    )

    # 2. 대기현황판(Board)에게는 설정이 켜져 있을 때만 전송
    settings = db.query(StoreSettings).filter(StoreSettings.store_id == current_store.id).first()
    should_broadcast_board = True
    if settings:
        use_board = getattr(settings, 'enable_waiting_board', True) 
        use_desk = getattr(settings, 'enable_reception_desk', True)
        if not use_board and not use_desk:
            should_broadcast_board = False
            logger.debug(f"Skipping member_updated broadcast to BOARD/RECEPTION (Board: disabled, Reception: disabled): store_id={current_store.id}")

    if should_broadcast_board:
        await sse_manager.broadcast(
            store_id=str(current_store.id),
            event_type="member_updated",
            data={
                "member_id": db_member.id,
                "name": db_member.name,
                "phone": db_member.phone
            },
            franchise_id=None,
            target_role='board'
        )

    # 핸드폰 번호가 변경되었거나, 기존 대기 내역에 member_id가 없는 경우 동기화
    # (단순 이름 변경 시에도 기존 waiting list에 member_id가 연결되어 있어야 함)
    active_waitings = db.query(WaitingList).filter(
        WaitingList.phone == db_member.phone,
        WaitingList.status == "waiting",
        WaitingList.store_id == current_store.id
    ).all()

    for w in active_waitings:
        if w.member_id != db_member.id:
            w.member_id = db_member.id
    
    if active_waitings:
        db.commit()

    return db_member

@router.delete("/{member_id}")
async def delete_member(
    member_id: int,
    current_store: Store = Depends(get_current_store),
    db: Session = Depends(get_db)
):
    """회원 삭제"""
    db_member = db.query(Member).filter(
        Member.id == member_id,
        Member.store_id == current_store.id
    ).first()

    if not db_member:
        raise HTTPException(status_code=404, detail="회원을 찾을 수 없습니다.")

    db.delete(db_member)
    db.commit()

    return {"message": "회원이 삭제되었습니다."}

@router.post("/bulk")
async def bulk_create_members(
    members: MemberBulkCreate,
    current_store: Store = Depends(get_current_store),
    db: Session = Depends(get_db)
):
    """회원 일괄 등록"""
    success_count = 0
    error_count = 0
    errors = []

    processed_phones = set()

    for member_data in members.members:
        try:
            # 배치 내 중복 확인
            if member_data.phone in processed_phones:
                error_count += 1
                errors.append({
                    "name": member_data.name,
                    "phone": member_data.phone,
                    "error": "목록 내 중복된 핸드폰번호"
                })
                continue

            # DB 중복 확인
            conflict_type, existing = check_member_uniqueness(db, current_store, phone=member_data.phone, barcode=member_data.barcode)
            if existing:
                error_count += 1
                if conflict_type == "barcode":
                     msg = "이미 등록된 바코드"
                else:
                    msg = "이미 등록된 핸드폰번호"
                    if current_store.franchise and current_store.franchise.member_type == "franchise":
                        msg += " (프랜차이즈 통합)"
                
                errors.append({
                    "name": member_data.name,
                    "phone": member_data.phone,
                    "error": msg
                })
                continue

            # 회원 등록
            db_member = Member(**member_data.dict(), store_id=current_store.id)
            db.add(db_member)
            processed_phones.add(member_data.phone)
            success_count += 1

        except Exception as e:
            error_count += 1
            errors.append({
                "name": member_data.name,
                "phone": member_data.phone,
                "error": str(e)
            })

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"저장 중 오류가 발생했습니다: {str(e)}")

    return {
        "message": f"총 {success_count}명 등록, {error_count}명 실패",
        "success_count": success_count,
        "error_count": error_count,
        "errors": errors
    }

@router.get("/sample-excel")
async def get_sample_excel():
    """회원 일괄 등록용 샘플 엑셀 다운로드"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "회원 일괄 등록"
    
    # 헤더
    ws.append(["고객명", "핸드폰번호", "바코드"])
    
    # 샘플 데이터
    ws.append(["홍길동", "01012345678", "BC123456"])
    ws.append(["김철수", "01087654321", ""])
    
    # 엑셀 파일 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=member_sample.xlsx"}
    )

@router.get("/statistics/ranking")
async def get_member_attendance_ranking(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    limit: int = 10,
    current_store: Store = Depends(get_current_store),
    db: Session = Depends(get_db)
):
    """
    회원 출석 순위 조회 (기본값: 이번달)
    """
    if not start_date:
        today = date.today()
        start_date = date(today.year, today.month, 1)
    if not end_date:
        end_date = date.today()

    # WaitingList(attended) + Member Join
    results = db.query(
        Member.id,
        Member.name,
        Member.phone,
        func.count(WaitingList.id).label("visit_count"),
        func.max(WaitingList.attended_at).label("last_visit")
    ).join(
        WaitingList, Member.id == WaitingList.member_id
    ).filter(
        WaitingList.store_id == current_store.id,
        WaitingList.status == "attended",
        WaitingList.business_date >= start_date,
        WaitingList.business_date <= end_date
    ).group_by(
        Member.id
    ).order_by(
        desc("visit_count")
    ).limit(limit).all()

    return [
        {
            "id": r.id,
            "name": r.name,
            "phone": r.phone,
            "visit_count": r.visit_count,
            "last_visit": r.last_visit
        }
        for r in results
    ]

@router.get("/statistics/new")
async def get_new_members_statistics(
    period: str = "month", # today, week, month
    current_store: Store = Depends(get_current_store),
    db: Session = Depends(get_db)
):
    """
    신규 회원 목록 조회
    """
    today = date.today()
    start_date = today

    if period == "week":
        start_date = today - timedelta(days=today.weekday()) # This Monday
    elif period == "month":
        start_date = date(today.year, today.month, 1)
    
    # Filter by created_at (Member registration)
    # Note: Member.created_at is DateTime, so we compare with date
    query = db.query(Member).filter(
        Member.store_id == current_store.id,
        Member.created_at >= datetime.combine(start_date, datetime.min.time())
    ).order_by(desc(Member.created_at))

    results = query.all()

    return [
        {
            "id": r.id,
            "name": r.name,
            "phone": r.phone,
            "created_at": r.created_at,
            "visit_count": 0 # Placeholder or separate query if needed
        }
        for r in results
    ]

@router.get("/statistics/inactive")
async def get_inactive_members(
    threshold_days: int = 30,
    current_store: Store = Depends(get_current_store),
    db: Session = Depends(get_db)
):
    """
    장기 미출석 회원 조회 (최근 방문이 threshold_days 이전인 회원)
    """
    today = datetime.now()
    threshold_date = today - timedelta(days=threshold_days)

    # Subquery: Get last visit date for each member
    last_visit_subquery = db.query(
        WaitingList.member_id,
        func.max(WaitingList.attended_at).label("last_visit")
    ).filter(
        WaitingList.store_id == current_store.id,
        WaitingList.status == "attended"
    ).group_by(
        WaitingList.member_id
    ).subquery()

    # Members who have visited before but not recently
    # Join Member with Subquery
    results = db.query(
        Member,
        last_visit_subquery.c.last_visit
    ).join(
        last_visit_subquery, Member.id == last_visit_subquery.c.member_id
    ).filter(
        Member.store_id == current_store.id,
        last_visit_subquery.c.last_visit < threshold_date
    ).order_by(
        last_visit_subquery.c.last_visit.asc() # Longest time since visit first
    ).limit(50).all()

    return [
        {
            "id": member.id,
            "name": member.name,
            "phone": member.phone,
            "last_visit": last_visit,
            "days_since": (today - last_visit).days
        }
        for member, last_visit in results
    ]
async def upload_excel(
    file: UploadFile = File(...),
    current_store: Store = Depends(get_current_store),
    db: Session = Depends(get_db)
):
    """
    엑셀 파일 업로드 및 검수
    - 엑셀 파일을 읽어서 회원 데이터 추출
    - 유효성 검사 후 등록 가능한 목록 반환
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="엑셀 파일만 업로드 가능합니다.")

    try:
        # 엑셀 파일 읽기
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents))
        sheet = workbook.active

        valid_members = []
        invalid_members = []
        processed_phones = set()

        # 첫 번째 행은 헤더로 간주하고 스킵
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 2:
                continue

            name = str(row[0]).strip() if row[0] else ""
            # 전화번호는 숫자로 읽혀서 변환되는 경우 처리
            phone_raw_value = row[1]
            if phone_raw_value is None:
                phone_raw = ""
            elif isinstance(phone_raw_value, (int, float)):
                # 숫자로 읽힌 경우 (예: 10이 "010-"로 입력된 경우)
                phone_raw = str(int(phone_raw_value))
                # 10, 100, 1000 등의 숫자는 010으로 시작하는 것으로 간주하고 앞에 0을 붙임
                if len(phone_raw) < 11 and phone_raw.startswith('10'):
                    phone_raw = '0' + phone_raw
            else:
                phone_raw = str(phone_raw_value).strip()

            # 하이픈 제거 (010-0000-0000 형식도 허용)
            phone = phone_raw.replace('-', '').replace(' ', '')

            # 유효성 검사
            errors = []

            if not name:
                errors.append("이름 없음")

            if not phone_raw:
                errors.append("핸드폰번호 없음")
            elif not phone.startswith("010") or len(phone) != 11 or not phone.isdigit():
                errors.append("핸드폰번호 형식 오류 (010-0000-0000 또는 01000000000)")

            # 바코드 읽기 (3번째 열, 옵션)
            barcode = None
            if len(row) > 2 and row[2]:
                barcode = str(row[2]).strip()

            # 중복 확인 (매장별)
            if phone:
                if phone in processed_phones:
                    errors.append("파일 내 중복된 번호")
                else:
                    conflict_type, existing = check_member_uniqueness(db, current_store, phone=phone, barcode=barcode)
                    if existing:
                        if conflict_type == "barcode":
                            errors.append("이미 등록된 바코드")
                        else:
                            msg = "이미 등록된 번호"
                            if current_store.franchise and current_store.franchise.member_type == "franchise":
                                msg += " (프랜차이즈 통합)"
                            errors.append(msg)
                    else:
                        processed_phones.add(phone)

            if errors:
                invalid_members.append({
                    "row": row_idx,
                    "name": name,
                    "phone": phone,
                    "errors": errors
                })
            else:
                valid_members.append({
                    "name": name,
                    "phone": phone
                })

        return {
            "total_count": len(valid_members) + len(invalid_members),
            "valid_count": len(valid_members),
            "invalid_count": len(invalid_members),
            "valid_members": valid_members,
            "invalid_members": invalid_members
        }

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"엑셀 파일 처리 중 오류: {str(e)}")

@router.post("/confirm-excel")
async def confirm_excel_upload(
    members: MemberBulkCreate,
    current_store: Store = Depends(get_current_store),
    db: Session = Depends(get_db)
):
    """
    엑셀 검수 후 최종 등록
    """
    return await bulk_create_members(members, current_store, db)

from schemas import QuickRegisterRequest, WaitingListCreate, QuickRegisterResponse
from routers.waiting import register_waiting

@router.post("/quick-register", response_model=QuickRegisterResponse)
async def quick_register(
    req: QuickRegisterRequest,
    current_store: Store = Depends(get_current_store),
    db: Session = Depends(get_db)
):
    """
    간편 등록 (관리자용)
    - 입력값(input_value)이 핸드폰번호 뒷자리(4자리) 또는 전체 번호, 또는 바코드일 수 있음
    - 회원을 찾아서 대기 등록
    """
    input_val = req.input_value.strip()
    target_member = None
    
    # 1. 바코드로 검색
    target_member = db.query(Member).filter(
        Member.barcode == input_val,
        Member.store_id == current_store.id
    ).first()
    
    # 2. 핸드폰 번호로 검색
    if not target_member:
        # 숫자만 추출
        digits = ''.join(filter(str.isdigit, input_val))
        
        if len(digits) == 11:
            # 11자리: 전체 번호 (01012345678)
             target_member = db.query(Member).filter(
                Member.phone == digits,
                Member.store_id == current_store.id
            ).first()

        elif len(digits) == 8:
            # 8자리: 010 + 8자리로 간주
            full_phone = "010" + digits
            target_member = db.query(Member).filter(
                Member.phone == full_phone,
                Member.store_id == current_store.id
            ).first()
            if not target_member:
                # 회원이 아니면 이 번호로 등록 진행
                pass

        elif len(digits) == 4:
            # 뒷자리 검색
            candidates = db.query(Member).filter(
                Member.phone.endswith(digits),
                Member.store_id == current_store.id
            ).all()
            
            if len(candidates) == 1:
                target_member = candidates[0]
            elif len(candidates) > 1:
                # 중복 발생: 후보 리스트 반환
                return QuickRegisterResponse(
                    success=False,
                    message="중복된 회원이 존재합니다.",
                    candidates=candidates
                )
            else:
                 raise HTTPException(status_code=404, detail="해당 번호의 회원을 찾을 수 없습니다.")
        else:
             # 그 외 길이는 일단 바코드나 잘못된 입력으로 간주 -> 위에서 바코드 없었으므로 에러
             raise HTTPException(status_code=400, detail="올바른 형식의 번호나 바코드를 입력해주세요.")

    # 대기 등록 준비
    phone_to_register = ""
    name_to_register = ""
    
    if target_member:
        phone_to_register = target_member.phone
        name_to_register = target_member.name
    else:
        # 멤버가 없는 경우 (전체 번호 입력 시)
        digits = ''.join(filter(str.isdigit, input_val))
        if len(digits) == 11 and digits.startswith("010"):
            phone_to_register = digits
            name_to_register = "" # 이름 없음
        elif len(digits) == 8:
             phone_to_register = "010" + digits
             name_to_register = ""
        else:
             raise HTTPException(status_code=404, detail="회원을 찾을 수 없으며 올바른 핸드폰 번호 형식이 아닙니다.")
            
    # 대기 등록 함수 호출
    waiting_create = WaitingListCreate(
        phone=phone_to_register,
        name=name_to_register,
        class_id=req.class_id,
        person_count=req.person_count,
        is_admin_registration=True
    )
    
    result = await register_waiting(waiting_create, db, current_store)
    
    return QuickRegisterResponse(
        success=True,
        data=result
    )
